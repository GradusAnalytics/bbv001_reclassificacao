"""
io_utils.py — Read inputs and write outputs for the BBV001 pipeline.

One function per Alteryx Input tool. Reading and the trivial first Select
(rename / column-type change) are bundled together where it makes sense.
"""
import logging
from pathlib import Path
from typing import Optional

import pandas as pd

from config import (
    INPUT_FILES,
    MANUAL_GROUP_OVERRIDES,
    OUTPUT_RECLASSIFIER,
    OUTPUT_FINAL_CONSOLIDATED,
    OUTPUT_DIR,
)
from helpers import log_step

logger = logging.getLogger(__name__)

# Assinatura do container binário OLE2/CFB usado por formatos legados do Office
# (.xls, .xlsb). main() sempre grava os inputs com extensão .xlsx fixa (ver CANON em
# bbv001_reclassificacao.py), então a extensão do arquivo não é confiável para saber
# se o conteúdo é realmente .xlsb — é preciso inspecionar os bytes.
_OLE2_SIGNATURE = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _excel_engine_for(path) -> str | None:
    """Detecta pelos magic bytes se o arquivo é um binário OLE2/CFB (.xlsb) e, se
    for, indica o engine 'pyxlsb'. Caso contrário, devolve None (engine padrão do
    pandas, que já lida bem com .xlsx reais)."""
    try:
        with open(path, "rb") as fh:
            header = fh.read(8)
    except OSError:
        return None
    return "pyxlsb" if header.startswith(_OLE2_SIGNATURE) else None


def _read_excel(path, **kwargs):
    """Wrapper de pd.read_excel que detecta .xlsb pelo conteúdo (não pela extensão)
    e seleciona o engine correto automaticamente."""
    engine = _excel_engine_for(path)
    if engine:
        kwargs = {**kwargs, "engine": engine}
    return pd.read_excel(path, **kwargs)


def _read_excel_with_header_marker(path, sheet, marker: str, max_scan: int = 15) -> pd.DataFrame:
    """
    Some sheets carry title/banner rows above the real table header (e.g. the
    'Base' and 'Unico CV' sheets start with 'Titulo2'/'Titulo3' and blank rows, so a
    plain read_excel mislabels the columns as 'Unnamed: N'). Scan the first `max_scan`
    rows for the cell equal to `marker` and use that row as the header.
    """
    probe = _read_excel(path, sheet_name=sheet, header=None, nrows=max_scan)
    header_row = None
    for i in range(len(probe)):
        if probe.iloc[i].astype(str).str.strip().eq(marker).any():
            header_row = i
            break
    if header_row is None:
        raise ValueError(
            f"Header marker {marker!r} not found in the first {max_scan} rows of "
            f"sheet {sheet!r} ({path}). The template may have changed."
        )
    return _read_excel(path, sheet_name=sheet, header=header_row)


# -------------------------------------------------------------------------
# Inputs
# -------------------------------------------------------------------------
def read_base_fechamento() -> pd.DataFrame:
    """Tool 4 + Tool 87 (rename Valor do Lancamento → Valor)."""
    cfg = INPUT_FILES["base_fechamento"]
    df = _read_excel(cfg["path"], sheet_name=cfg["sheet"])
    df = df.rename(columns={"Valor do Lancamento": "Valor"})
    # Conta Contabil cells come back as full-precision Python ints (25-digit account codes
    # stored as numbers in Excel). Keep them as strings: as ints they survive the pipeline
    # in memory but get cast to float64 on the Excel write, losing precision
    # (8172700000000000010000000 → 8172699999999999715835904). This column also feeds
    # 'Conta destino' for the two Match paths (Tools 93/94), so the fix covers both.
    df["Conta Contabil"] = df["Conta Contabil"].apply(
        lambda v: v if (isinstance(v, str) or (isinstance(v, float) and pd.isna(v))) else str(v)
    )
    log_step(logger, "4", "Read Base Fechamento + rename Valor", df)
    return df


def read_depara_custo() -> pd.DataFrame:
    """Tool 10."""
    cfg = INPUT_FILES["depara_custo"]
    df = _read_excel(cfg["path"], sheet_name=cfg["sheet"])
    log_step(logger, "10", "Read De-Para Custo", df)
    return df


def read_classe_valor_conta() -> pd.DataFrame:
    """Tool 47 — sheet 'Base' (real header below a few title rows)."""
    cfg = INPUT_FILES["classe_valor_conta"]
    df = _read_excel_with_header_marker(cfg["path"], cfg["sheet_base"], "Nome classe de valor")
    log_step(logger, "47", "Read Classe Valor x Conta (Base)", df)
    return df


def read_unico_cv() -> pd.DataFrame:
    """Tool 48 — sheet 'Unico CV' (real header below a few title rows)."""
    cfg = INPUT_FILES["classe_valor_conta"]
    df = _read_excel_with_header_marker(cfg["path"], cfg["sheet_unico_cv"], "Classe de valor")
    log_step(logger, "48", "Read Unico CV", df)
    return df


def read_estrutura_contas() -> pd.DataFrame:
    """Tool 61."""
    cfg = INPUT_FILES["estrutura_contas"]
    df = _read_excel(cfg["path"], sheet_name=cfg["sheet"])
    log_step(logger, "61", "Read Estrutura de Contas", df)
    return df


def read_base_reclassificada() -> Optional[pd.DataFrame]:
    """
    Tool 124. Returns None when the file does not exist —
    e.g., on the first run that GENERATES the reclassifier input.
    On the second run (which CONSUMES the reclassifier output), the file must exist.
    """
    cfg = INPUT_FILES["base_reclassificada"]
    path: Path = cfg["path"]
    if not path.exists():
        logger.warning(
            f"[Tool 124] base_reclassificada not found at {path}. "
            f"Assuming 1st-run mode (no reclassifier output yet). "
            f"Reclassificador path will return an empty frame."
        )
        return None
    df = _read_excel(path, sheet_name=cfg["sheet"])
    log_step(logger, "124", "Read base_reclassificada", df)
    return df


def read_estrutura_entidades_cc() -> Optional[pd.DataFrame]:
    """
    NOVO (Tool 200) — cadastro de Entidades x Centros de Custo.

    Input OPCIONAL e tolerante a falhas: qualquer problema de leitura apenas emite
    WARNING e retorna None (a aba 'Centros de Custo' do relatório fica vazia) — nunca
    derruba a execução, pois as demais saídas não dependem deste cadastro.

    .xlsx real: lido via openpyxl com data_only=True (pega o valor já calculado das
    fórmulas). .xlsb (detectado pelos magic bytes, não pela extensão): lido via
    pyxlsb — atenção que pyxlsb NÃO recalcula fórmulas, só devolve o valor bruto
    armazenado no arquivo; se o cadastro usar fórmulas, confira se os valores saíram
    corretos (emite um WARNING adicional nesse caso).
    """
    cfg = INPUT_FILES["estrutura_entidades_cc"]
    path: Path = cfg["path"]
    if not path.exists():
        logger.warning(
            f"[Tool 200] cadastro de Centros de Custo não encontrado em {path}. "
            f"A aba 'Centros de Custo' do relatório ficará vazia."
        )
        return None
    try:
        if _excel_engine_for(path) == "pyxlsb":
            logger.warning(
                "[Tool 200] cadastro de CC recebido em .xlsb — pyxlsb não recalcula "
                "fórmulas (só openpyxl com data_only=True faz isso); se o cadastro "
                "depender de fórmulas, confira se os valores saíram corretos."
            )
            xls = pd.ExcelFile(path, engine="pyxlsb")
            sheet = cfg["sheet"] if cfg["sheet"] in xls.sheet_names else xls.sheet_names[0]
            df = xls.parse(sheet)
        else:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            sheet = cfg["sheet"] if cfg["sheet"] in wb.sheetnames else wb.sheetnames[0]
            rows = list(wb[sheet].values)
            wb.close()
            if not rows:
                logger.warning("[Tool 200] cadastro de CC vazio. Aba 'Centros de Custo' ficará vazia.")
                return None
            header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
            df = pd.DataFrame(rows[1:], columns=header)
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            f"[Tool 200] não foi possível ler o cadastro de CC ({exc}). "
            f"A aba 'Centros de Custo' ficará vazia."
        )
        return None
    log_step(logger, "200", "Read Estrutura de Entidades x CC", df)
    return df


def get_manual_overrides() -> pd.DataFrame:
    """Tool 86 — hardcoded TextInput, 11 rows."""
    df = pd.DataFrame(MANUAL_GROUP_OVERRIDES)
    log_step(logger, "86", "Manual Grupo overrides (TextInput)", df)
    return df


# -------------------------------------------------------------------------
# Outputs
# -------------------------------------------------------------------------
def write_reclassifier_base(df: pd.DataFrame) -> Path:
    """Tool 123 — base para o processo de reclassificação."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUTPUT_RECLASSIFIER, sheet_name="Base", index=False)
    log_step(logger, "123", f"Wrote reclassifier base → {OUTPUT_RECLASSIFIER.name}", df)
    return OUTPUT_RECLASSIFIER


def write_final_consolidated(df: pd.DataFrame) -> Path:
    """Tool 114 equivalent — the consolidated base for matrix upload.
    (Alteryx workflow did not write this; we add it because user confirmed it's needed.)
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_excel(OUTPUT_FINAL_CONSOLIDATED, sheet_name="Consolidado", index=False)
    log_step(logger, "114", f"Wrote final consolidated → {OUTPUT_FINAL_CONSOLIDATED.name}", df)
    return OUTPUT_FINAL_CONSOLIDATED
